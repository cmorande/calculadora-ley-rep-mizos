"""Lee archivos de declaración ya calculados (generados por esta app, o
declaraciones oficiales ya presentadas que respeten el mismo layout de
celdas) y arma un DataFrame largo para comparar variación entre periodos."""

from __future__ import annotations

import re

import pandas as pd
from openpyxl import load_workbook

from . import taxonomia
from .exportar import COLS_DOMICILIARIO, COLS_NO_DOMICILIARIO


class ErrorDeclaracion(Exception):
    pass


def sugerir_periodo(nombre_archivo: str) -> str:
    nombre = re.sub(r"\.xlsx?$", "", nombre_archivo, flags=re.IGNORECASE)
    nombre = re.sub(r"(?i)declaraci[oó]n\s*", "", nombre).strip(" _-")
    return nombre.title() if nombre else nombre_archivo


def leer_declaracion(fuente, periodo: str, sheet_name: str = "LB") -> pd.DataFrame:
    try:
        wb = load_workbook(fuente, data_only=True)
    except Exception as e:
        raise ErrorDeclaracion(f"No se pudo abrir el archivo: {e}") from e

    hoja = sheet_name if sheet_name in wb.sheetnames else wb.sheetnames[0]
    ws = wb[hoja]

    filas = []
    bloques = [
        ("Domiciliario", COLS_DOMICILIARIO),
        ("No Domiciliario", COLS_NO_DOMICILIARIO),
    ]
    for categoria, cols in bloques:
        bloque_tax = taxonomia.BLOQUES[categoria]
        for i, ft in enumerate(bloque_tax):
            fila = 7 + i
            no_pel = ws.cell(row=fila, column=cols["no_pel"]).value
            pel = ws.cell(row=fila, column=cols["pel"]).value
            filas.append(
                {
                    "Periodo": periodo,
                    "Categoría": categoria,
                    "Subcategoría": ft.subcategoria,
                    "Subcategoria2": ft.subcategoria2,
                    "Material": ft.material,
                    "No Peligroso": float(no_pel) if isinstance(no_pel, (int, float)) else 0.0,
                    "Peligroso": float(pel) if isinstance(pel, (int, float)) else 0.0,
                }
            )

    df = pd.DataFrame(filas)
    df["Toneladas"] = df["No Peligroso"] + df["Peligroso"]

    if df["Toneladas"].sum() == 0:
        raise ErrorDeclaracion(
            "El archivo se leyó pero todas las toneladas son 0: revisa que sea una "
            "declaración con el formato RESIMPLE (hoja 'LB', mismas filas/columnas)."
        )
    return df


def combinar(dataframes: list[pd.DataFrame]) -> pd.DataFrame:
    return pd.concat(dataframes, ignore_index=True)


def resumen_por_periodo(df_largo: pd.DataFrame) -> pd.DataFrame:
    """Toneladas totales por Periodo x Categoría (Domiciliario/No Domiciliario)."""
    return df_largo.groupby(["Periodo", "Categoría"], as_index=False)["Toneladas"].sum()


def composicion_por_periodo(df_largo: pd.DataFrame) -> pd.DataFrame:
    """Toneladas por Periodo x Subcategoría principal, sumando ambas categorías."""
    return df_largo.groupby(["Periodo", "Subcategoría"], as_index=False)["Toneladas"].sum()


def tabla_variacion(df_largo: pd.DataFrame, orden_periodos: list[str]) -> pd.DataFrame:
    """Pivotea Periodo (columnas, en el orden dado) x Subcategoría (filas) con
    toneladas totales, y agrega columnas de variación % entre periodos consecutivos."""
    comp = composicion_por_periodo(df_largo)
    pivot = comp.pivot(index="Subcategoría", columns="Periodo", values="Toneladas").fillna(0.0)
    pivot = pivot.reindex(columns=[p for p in orden_periodos if p in pivot.columns])
    pivot = pivot.reindex(index=taxonomia.SUBCATEGORIAS_ORDEN).fillna(0.0)

    columnas_periodo = list(pivot.columns)
    for i in range(1, len(columnas_periodo)):
        anterior, actual = columnas_periodo[i - 1], columnas_periodo[i]
        col_var = f"Var. {anterior} → {actual} (%)"
        variacion = ((pivot[actual] - pivot[anterior]) / pivot[anterior].replace(0, pd.NA)) * 100
        pivot[col_var] = variacion.replace([float("inf"), float("-inf")], pd.NA)
    return pivot.round(3)

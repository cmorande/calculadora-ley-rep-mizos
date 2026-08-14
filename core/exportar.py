"""Genera el archivo .xlsx de declaración de envases y embalajes en el
formato oficial RESIMPLE (hoja "LB"), a partir de los resultados de
core.calculo.calcular().

El layout (filas, columnas y celdas combinadas) replica exactamente una
declaración real presentada, para maximizar compatibilidad con lo que
espera revisar/cargar la empresa.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import taxonomia
from .calculo import ResultadoCalculo

# Columnas (1-based) por bloque: (subcategoria, subcategoria2/material2, material, no_peligroso, peligroso)
COLS_DOMICILIARIO = dict(sub=2, sub2_o_mat=3, mat=4, no_pel=5, pel=6)  # B, C, D, E, F
COLS_NO_DOMICILIARIO = dict(sub=8, sub2_o_mat=9, mat=10, no_pel=11, pel=12)  # H, I, J, K, L

TITULO_FILL = PatternFill("solid", fgColor="DCE6F1")
HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")


@dataclass
class InfoEmpresa:
    razon_social: str = ""
    responsable: str = ""
    id_rut_reporta: str = ""
    rut_empresa: str = ""
    representante_legal: str = ""


def _escribir_encabezado(ws: Worksheet, info: InfoEmpresa) -> None:
    ws["B1"] = "Nombre Productor:"
    ws["C1"] = info.razon_social
    ws.merge_cells("C1:D1")
    ws["J1"] = "ID RUT de empresa que reporta"
    ws["K1"] = info.id_rut_reporta
    ws.merge_cells("K1:L1")

    ws["B2"] = "Responsable:"
    ws["C2"] = info.responsable
    ws.merge_cells("C2:D2")
    ws["J2"] = "RUT empresa"
    ws["K2"] = info.rut_empresa
    ws.merge_cells("K2:L2")

    ws["J3"] = "Representante Legal"
    ws["K3"] = info.representante_legal
    ws.merge_cells("K3:L3")

    for coord in ("B1", "B2", "J1", "J2", "J3"):
        ws[coord].font = BOLD

    ws["B5"] = "CATEGORIA DOMICILIARIA"
    ws.merge_cells("B5:F5")
    ws["H5"] = "CATEGORIA NO DOMICILIARIA"
    ws.merge_cells("H5:L5")
    for coord in ("B5", "H5"):
        ws[coord].font = BOLD
        ws[coord].alignment = CENTER
        ws[coord].fill = TITULO_FILL

    ws["B6"] = "SUB CATEGORIA"
    ws.merge_cells("B6:C6")
    ws["D6"] = "MATERIAL"
    ws["E6"] = "NO PELIGROSO (TONELADAS)"
    ws["F6"] = "PELIGROSOS (TONELADAS)"
    ws["H6"] = "SUB CATEGORIA"
    ws.merge_cells("H6:I6")
    ws["J6"] = "MATERIAL"
    ws["K6"] = "NO PELIGROSO (TONELADAS)"
    ws["L6"] = "PELIGROSOS (TONELADAS)"
    for coord in ("B6", "D6", "E6", "F6", "H6", "J6", "K6", "L6"):
        ws[coord].font = BOLD
        ws[coord].alignment = CENTER
        ws[coord].fill = HEADER_FILL


def _merge_contiguo(ws: Worksheet, col: int, fila_ini: int, fila_fin: int) -> None:
    if fila_fin > fila_ini:
        letra = get_column_letter(col)
        ws.merge_cells(f"{letra}{fila_ini}:{letra}{fila_fin}")


def _escribir_bloque(ws: Worksheet, categoria: str, cols: dict, valores: list[list[float]], fila_inicio: int) -> None:
    filas_tax = taxonomia.BLOQUES[categoria]
    n = len(filas_tax)

    # Rangos contiguos de subcategoría (col "sub") para fusionar
    inicio_sub = fila_inicio
    for i in range(n):
        fila = fila_inicio + i
        actual = filas_tax[i]
        es_ultimo = i == n - 1
        distinto_siguiente = es_ultimo or filas_tax[i + 1].subcategoria != actual.subcategoria
        if i == 0 or filas_tax[i - 1].subcategoria != actual.subcategoria:
            ws.cell(row=fila, column=cols["sub"], value=actual.subcategoria).font = BOLD
            inicio_sub = fila
        if distinto_siguiente:
            _merge_contiguo(ws, cols["sub"], inicio_sub, fila)

    # Rangos contiguos de subcategoría2 (Flexible/Rígido) para fusionar, y material
    inicio_sub2 = fila_inicio
    for i in range(n):
        fila = fila_inicio + i
        actual = filas_tax[i]
        if actual.subcategoria2 is None:
            # Sin split: el texto del material ocupa la celda "sub2_o_mat" fusionada con "mat"
            ws.cell(row=fila, column=cols["sub2_o_mat"], value=actual.material)
            ws.merge_cells(
                start_row=fila, start_column=cols["sub2_o_mat"], end_row=fila, end_column=cols["mat"]
            )
        else:
            es_ultimo = i == n - 1
            distinto_siguiente = (
                es_ultimo
                or filas_tax[i + 1].subcategoria2 != actual.subcategoria2
                or filas_tax[i + 1].subcategoria != actual.subcategoria
            )
            if i == 0 or (
                filas_tax[i - 1].subcategoria2 != actual.subcategoria2
                or filas_tax[i - 1].subcategoria != actual.subcategoria
            ):
                ws.cell(row=fila, column=cols["sub2_o_mat"], value=actual.subcategoria2)
                inicio_sub2 = fila
            if distinto_siguiente:
                _merge_contiguo(ws, cols["sub2_o_mat"], inicio_sub2, fila)
            ws.cell(row=fila, column=cols["mat"], value=actual.material)

        no_pel, pel = valores[i]
        c_no_pel = ws.cell(row=fila, column=cols["no_pel"], value=round(no_pel, 4))
        c_pel = ws.cell(row=fila, column=cols["pel"], value=round(pel, 4))
        c_no_pel.number_format = "0.0000"
        c_pel.number_format = "0.0000"


def generar_workbook(resultado: ResultadoCalculo, info: InfoEmpresa) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "LB"

    _escribir_encabezado(ws, info)
    _escribir_bloque(ws, "Domiciliario", COLS_DOMICILIARIO, resultado.valores["Domiciliario"], fila_inicio=7)
    _escribir_bloque(ws, "No Domiciliario", COLS_NO_DOMICILIARIO, resultado.valores["No Domiciliario"], fila_inicio=7)

    anchos = {"B": 22, "C": 16, "D": 42, "E": 14, "F": 14, "H": 16, "I": 16, "J": 42, "K": 14, "L": 14}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    return wb


def generar_bytes(resultado: ResultadoCalculo, info: InfoEmpresa) -> bytes:
    wb = generar_workbook(resultado, info)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
